import os
import re
import json
import time
import logging
import argparse
import pandas as pd
from typing import List
from pydantic import BaseModel, Field, validator
from dotenv import load_dotenv
from constants import ACCOUNT_ID
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import PydanticOutputParser
from mask import mask_text
from notifications.database_config import PostgresConfig

load_dotenv()
os.makedirs("logs", exist_ok=True)

# ── LOGGING ───────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(name)s | %(message)s',
    handlers=[
        logging.FileHandler('logs/clan_metadata.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('metadata')

# ── SETTINGS ──────────────────────────────────────────
CATALOG_FILE   = "./video_catalog_with_content.xlsx"
OUTPUT_FILE    = "./video_catalog_enriched.xlsx"
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
MAX_RETRIES    = 3
# ──────────────────────────────────────────────────────

# ── VALID VALUES (Dynamic) ────────────────────────────
VALID_INDICATORS = []
_VALID_INDICATORS_LOADED = False


def load_valid_indicators() -> list[str]:
    global VALID_INDICATORS, _VALID_INDICATORS_LOADED
    if _VALID_INDICATORS_LOADED:
        return VALID_INDICATORS

    try:
        db_indicators = PostgresConfig(account_id=ACCOUNT_ID).get_indicators()
        codes = [str(ind.get("code", "")).strip().lower().replace(" ", "_") for ind in db_indicators]
        codes = [code for code in codes if code]
        if not codes:
            raise RuntimeError(f"No active KIIs found for account_id={ACCOUNT_ID}")
        VALID_INDICATORS = codes
        logger.info("Loaded %s indicators from PostgreSQL for account_id=%s", len(VALID_INDICATORS), ACCOUNT_ID)
    except Exception as exc:
        logger.error("Failed to load indicators from DB: %s", exc)
        raise

    _VALID_INDICATORS_LOADED = True
    return VALID_INDICATORS

VALID_AUDIENCES   = ["low_performer", "mid_performer", "top_performer", "all"]
VALID_DIFFICULTY  = ["beginner", "intermediate", "advanced"]
VALID_PHASES      = ["acquisition", "development", "conversion", "all"]
VALID_EXPERIENCE  = ["new_joiner", "experienced", "senior", "all"]
VALID_TONES       = ["supportive", "respectful", "encouraging", "calm", "neutral"]

GENERIC_SUMMARY_PATTERNS = [
    r"\bthis video (is|was) about\b",
    r"\bthis training video\b",
    r"\bprovides valuable insights\b",
    r"\blearn about\b",
    r"\boverview of\b",
]

HARSH_TONE_PATTERNS = [
    r"\byou failed\b",
    r"\byou are wrong\b",
    r"\bcareless\b",
    r"\blazy\b",
    r"\bstupid\b",
    r"\bincompetent\b",
    r"\bshame\b",
]

ROLE_RELEVANCE_HINTS = [
    "customer",
    "relationship manager",
    "rm",
    "lead",
    "loan",
    "sales",
    "huddle",
    "disbursal",
]

# ── PYDANTIC SCHEMA ───────────────────────────────────
class VideoMetadata(BaseModel):
    summary: str = Field(
        description="3-4 sentence summary of what this video teaches"
    )
    lead_indicators: List[str] = Field(
        description="2-3 lead indicators this video addresses. Must be from valid list only."
    )
    target_audience: str = Field(
        description="Who benefits most: low_performer / mid_performer / top_performer / all"
    )
    difficulty: str = Field(
        description="Content complexity: beginner / intermediate / advanced"
    )
    key_lesson: str = Field(
        description="One sentence — the single most important takeaway from this video"
    )
    sales_phase: str = Field(
        description="Which sales phase this helps: acquisition / development / conversion / all"
    )
    experience_level: str = Field(
        description="Who this is for: new_joiner / experienced / senior / all"
    )
    problem_solved: str = Field(
        description="One sentence — exactly what problem or struggle this video solves"
    )
    persona_type: str = Field(
        description="Persona archetype for the learner, e.g., new_joiner_rm / experienced_rm / senior_rm / all"
    )
    user_context: str = Field(
        description="One sentence describing the user's immediate work context"
    )
    background_situation: str = Field(
        description="One sentence giving relevant background conditions and constraints"
    )
    emotional_tone: str = Field(
        description="Tone used in summary: supportive / respectful / encouraging / calm / neutral"
    )

    @validator('lead_indicators')
    def validate_indicators(cls, v):
        if not VALID_INDICATORS:
            raise ValueError(f"No active lead indicators available in DB for account_id={ACCOUNT_ID}")
        cleaned = []
        for indicator in v:
            ind = indicator.strip().lower().replace(' ', '_')
            if ind in VALID_INDICATORS:
                cleaned.append(ind)
            else:
                # Try to find closest match
                for valid in VALID_INDICATORS:
                    if any(word in valid for word in ind.split('_')):
                        cleaned.append(valid)
                        break
        if not cleaned:
            cleaned = VALID_INDICATORS[:1]
        return cleaned[:3]  # max 3

    @validator('target_audience')
    def validate_audience(cls, v):
        v = v.strip().lower().replace(' ', '_')
        return v if v in VALID_AUDIENCES else "all"

    @validator('difficulty')
    def validate_difficulty(cls, v):
        v = v.strip().lower()
        return v if v in VALID_DIFFICULTY else "beginner"

    @validator('sales_phase')
    def validate_phase(cls, v):
        v = v.strip().lower()
        return v if v in VALID_PHASES else "all"

    @validator('experience_level')
    def validate_experience(cls, v):
        v = v.strip().lower().replace(' ', '_').replace('-', '_')
        return v if v in VALID_EXPERIENCE else "all"

    @validator('persona_type')
    def validate_persona_type(cls, v):
        v = v.strip().lower().replace(' ', '_').replace('-', '_')
        return v if v else "all"

    @validator('user_context')
    def validate_user_context(cls, v):
        text = str(v).strip()
        return text if text else "User handling day-to-day customer interactions and lead follow-ups."

    @validator('background_situation')
    def validate_background_situation(cls, v):
        text = str(v).strip()
        return text if text else "Working under target pressure while balancing customer trust and process quality."

    @validator('emotional_tone')
    def validate_emotional_tone(cls, v):
        v = v.strip().lower()
        return v if v in VALID_TONES else "supportive"


def _is_generic_summary(text: str) -> bool:
    cleaned = str(text or "").strip().lower()
    if len(cleaned.split()) < 18:
        return True
    return any(re.search(pattern, cleaned) for pattern in GENERIC_SUMMARY_PATTERNS)


def _has_harsh_tone(text: str) -> bool:
    cleaned = str(text or "").strip().lower()
    return any(re.search(pattern, cleaned) for pattern in HARSH_TONE_PATTERNS)


def _is_role_relevant(summary: str, key_lesson: str, problem_solved: str, role_text: str) -> bool:
    blob = " ".join([summary or "", key_lesson or "", problem_solved or "", role_text or ""]).lower()
    return any(hint in blob for hint in ROLE_RELEVANCE_HINTS)


def validate_metadata_quality(metadata: VideoMetadata, role_text: str):
    if _is_generic_summary(metadata.summary):
        return False, "generic_summary"
    if _has_harsh_tone(metadata.summary) or _has_harsh_tone(metadata.problem_solved):
        return False, "harsh_tone"
    if metadata.emotional_tone not in VALID_TONES:
        return False, "invalid_emotional_tone"
    if not _is_role_relevant(metadata.summary, metadata.key_lesson, metadata.problem_solved, role_text):
        return False, "role_not_relevant"
    return True, "ok"


# ── LANGCHAIN SETUP ───────────────────────────────────
def build_chain(api_key):
    indicators = load_valid_indicators()

    llm = ChatGoogleGenerativeAI(
        model="gemini-2.0-flash",
        google_api_key=api_key,
        temperature=0.1
    )

    parser = PydanticOutputParser(pydantic_object=VideoMetadata)

    prompt = PromptTemplate(
        template="""
You are a metadata extractor for sales training videos at a housing finance company.

VIDEO TITLE: {title}

SCREEN TEXT (what appears on screen):
{screen_text}

AUDIO TRANSCRIPT (what the person says):
{transcript}

VALID LEAD INDICATORS (use ONLY these exact values):
{valid_indicators}

TASK: Extract structured metadata from this video.

Rules:
- summary: 3-4 sentences describing what this video teaches. Be specific, not generic.
- lead_indicators: pick 2 from the valid list that best match
- target_audience: who benefits most (low_performer/mid_performer/top_performer/all)
- difficulty: how complex is the content (beginner/intermediate/advanced)
- key_lesson: ONE sentence — the most important takeaway
- sales_phase: which phase this helps (acquisition/development/conversion/all)
- experience_level: who this is for (new_joiner/experienced/senior/all)
- problem_solved: ONE sentence — what specific problem this video solves
- persona_type: learner persona label in snake_case (example: new_joiner_rm / experienced_rm / senior_rm / all)
- user_context: ONE sentence about the user's immediate work context
- background_situation: ONE sentence about background constraints and environment
- emotional_tone: choose ONE from supportive/respectful/encouraging/calm/neutral

Strict style rules (mandatory):
- respectful language
- non-judgmental language
- actionable guidance orientation
- no shaming language
- no harsh tone
- role-relevant to sales/RM/customer interaction context

{format_instructions}

Return ONLY valid JSON. No explanation. No markdown.
""",
        input_variables=["title", "screen_text", "transcript"],
        partial_variables={
            "valid_indicators": "\n".join(indicators),
            "format_instructions": parser.get_format_instructions()
        }
    )

    return prompt | llm | parser


def generate_metadata_for_video(chain, row, attempt=1):
    """Generate metadata for one video with retry logic"""

    title       = str(row.get('Title', row.get('title', '')))
    screen_text = str(row.get('screen_text', ''))
    transcript  = str(row.get('transcript', ''))
    creator     = str(row.get('creator_name', ''))
    role_text   = str(row.get('creator_role', ''))

    # ── Mask sensitive data ──
    masked_screen, screen_summary = mask_text(screen_text, creator_name=creator)
    masked_transcript, trans_summary = mask_text(transcript, creator_name=creator)

    logger.info(f"Screen masking: {screen_summary}")
    logger.info(f"Transcript masking: {trans_summary}")

    # ── Truncate if too long (Gemini has token limits) ──
    if len(masked_transcript) > 3000:
        masked_transcript = masked_transcript[:3000] + "..."
        logger.info("Transcript truncated to 3000 chars")

    try:
        logger.info(f"Calling Gemini API (attempt {attempt}/{MAX_RETRIES})...")
        start = time.time()

        result = chain.invoke({
            "title":       title,
            "screen_text": masked_screen,
            "transcript":  masked_transcript,
        })

        elapsed = time.time() - start
        logger.info(f"Gemini responded in {elapsed:.1f}s")
        logger.info(f"lead_indicators: {result.lead_indicators}")
        logger.info(f"target_audience: {result.target_audience}")
        logger.info(f"difficulty: {result.difficulty}")
        logger.info(f"sales_phase: {result.sales_phase}")
        logger.info(f"experience_level: {result.experience_level}")
        logger.info(f"persona_type: {result.persona_type}")
        logger.info(f"emotional_tone: {result.emotional_tone}")

        is_valid, reason = validate_metadata_quality(result, role_text=role_text)
        if not is_valid:
            raise ValueError(f"metadata_quality_rejected:{reason}")

        return result

    except Exception as e:
        logger.error(f"Attempt {attempt} failed: {str(e)}")

        if attempt < MAX_RETRIES:
            wait = attempt * 5
            logger.warning(f"Retrying in {wait}s...")
            time.sleep(wait)
            return generate_metadata_for_video(chain, row, attempt + 1)
        else:
            logger.error(f"All {MAX_RETRIES} attempts failed. Using fallback.")
            return None


def build_fallback(row):
    """Fallback metadata if Gemini fails"""
    title = str(row.get('Title', row.get('title', 'Unknown')))

    catalog_fallback = [
        str(i).strip().lower().replace(' ', '_')
        for i in str(row.get('lead_indicator', '')).split(',')
        if str(i).strip()
    ]
    if catalog_fallback:
        fallback_indicators = catalog_fallback
    else:
        try:
            fallback_indicators = load_valid_indicators()[:1]
        except Exception:
            fallback_indicators = []

    return {
        "summary":          f"Training video: {title}",
        "lead_indicators":  fallback_indicators,
        "target_audience":  "all",
        "difficulty":       "beginner",
        "key_lesson":       f"Key insights from: {title}",
        "sales_phase":      "all",
        "experience_level": "all",
        "problem_solved":   f"Addresses challenges related to: {title}",
        "persona_type":     "all",
        "user_context":     "User handling day-to-day customer interactions and lead follow-ups.",
        "background_situation": "Working under target pressure while balancing customer trust and process quality.",
        "emotional_tone":   "supportive",
    }


def process_catalog(api_key, input_file=CATALOG_FILE, limit=None):
    logger.info("Reading video catalog...")
    df = pd.read_excel(input_file)
    df = df.fillna("")

    if limit is not None:
        df = df.head(limit)
        logger.info("Test mode: processing first %d videos", len(df))

    logger.info(f"Found {len(df)} videos to process")

    # Check if transcript column exists
    has_transcript = 'transcript' in df.columns
    has_screen     = 'screen_text' in df.columns
    logger.info(f"Has transcript column: {has_transcript}")
    logger.info(f"Has screen_text column: {has_screen}")

    # Build LangChain
    logger.info("Building LangChain pipeline...")
    chain = build_chain(api_key)
    logger.info("LangChain ready")

    # Process each video
    results = []
    excluded_count = 0

    for i, (_, row) in enumerate(df.iterrows(), 1):
        title = str(row.get('Title', row.get('title', f'Video {i}')))
        
        logger.info(f"{'='*50}")
        logger.info(f"[{i}/{len(df)}] {title}")

        metadata = generate_metadata_for_video(chain, row)

        if metadata:
            # Check if summary indicates intro video — skip if so
            summary_lower = str(metadata.summary).lower()
            if 'introduction' in summary_lower or 'intro video' in summary_lower or 'onboarding' in summary_lower:
                logger.info(f"[{i}/{len(df)}] SKIPPED (summary indicates intro): {title}")
                excluded_count += 1
            else:
                results.append({
                    **row.to_dict(),
                    "ai_summary":          metadata.summary,
                    "ai_lead_indicators":  json.dumps(metadata.lead_indicators),
                    "ai_target_audience":  metadata.target_audience,
                    "ai_difficulty":       metadata.difficulty,
                    "ai_key_lesson":       metadata.key_lesson,
                    "ai_sales_phase":      metadata.sales_phase,
                    "ai_experience_level": metadata.experience_level,
                    "ai_problem_solved":   metadata.problem_solved,
                    "ai_persona_type":     metadata.persona_type,
                    "ai_user_context":     metadata.user_context,
                    "ai_background_situation": metadata.background_situation,
                    "ai_emotional_tone":   metadata.emotional_tone,
                    "ai_generated":        True,
                })
                logger.info(f"[{i}/{len(df)}] DONE: {title}")
        else:
            fallback = build_fallback(row)
            # Check if fallback summary indicates intro video
            summary_lower = str(fallback["summary"]).lower()
            if 'introduction' in summary_lower or 'intro video' in summary_lower or 'onboarding' in summary_lower:
                logger.info(f"[{i}/{len(df)}] SKIPPED (summary indicates intro): {title}")
                excluded_count += 1
            else:
                results.append({
                    **row.to_dict(),
                    "ai_summary":          fallback["summary"],
                    "ai_lead_indicators":  json.dumps(fallback["lead_indicators"]),
                    "ai_target_audience":  fallback["target_audience"],
                    "ai_difficulty":       fallback["difficulty"],
                    "ai_key_lesson":       fallback["key_lesson"],
                    "ai_sales_phase":      fallback["sales_phase"],
                    "ai_experience_level": fallback["experience_level"],
                    "ai_problem_solved":   fallback["problem_solved"],
                    "ai_persona_type":     fallback["persona_type"],
                    "ai_user_context":     fallback["user_context"],
                    "ai_background_situation": fallback["background_situation"],
                    "ai_emotional_tone":   fallback["emotional_tone"],
                    "ai_generated":        False,
                })
                logger.warning(f"[{i}/{len(df)}] FALLBACK used: {title}")

        # Rate limit — Gemini free tier allows 15 requests/min
        time.sleep(4)

    logger.info(f"Processing complete | processed={len(results)} | excluded={excluded_count} intro videos")
    return pd.DataFrame(results)


def save_enriched_catalog(df, output_file=OUTPUT_FILE):
    logger.info(f"Saving enriched catalog to {output_file}...")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Video Catalog"

    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[cell.column_letter].width = 30

    ws.row_dimensions[1].height = 30

    for row, (_, data) in enumerate(df.iterrows(), 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=str(value))
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )
        ws.row_dimensions[row].height = 100

    ws.freeze_panes = "A2"
    wb.save(output_file)
    logger.info(f"Saved to {output_file}")


# ── RUN ───────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate AI metadata for videos")
    parser.add_argument("--input", default=CATALOG_FILE, help="Input xlsx file")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Output xlsx file")
    parser.add_argument("--limit", type=int, default=None, help="Process only first N videos")
    parser.add_argument("--api-key", default=None, help="Gemini API key (overrides env)")
    args = parser.parse_args()

    api_key = (args.api_key or GEMINI_API_KEY or "").strip()

    if not api_key:
        logger.error("No API key provided. Set GEMINI_API_KEY in .env or pass --api-key.")
        exit(1)

    logger.info("Starting metadata generation for all videos...")
    logger.info(f"Input file: {args.input}")
    logger.info(f"Output will be saved to: {args.output}")

    df = process_catalog(api_key, input_file=args.input, limit=args.limit)
    save_enriched_catalog(df, output_file=args.output)

    logger.info("=" * 50)
    logger.info("METADATA GENERATION COMPLETE")
    logger.info(f"Total videos processed: {len(df)}")
    logger.info(f"AI generated: {df['ai_generated'].sum()}")
    logger.info(f"Fallback used: {(~df['ai_generated']).sum()}")
    logger.info(f"Output: {args.output}")